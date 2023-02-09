import inspect
import logging
import csv
import softest
from openpyxl import Workbook, load_workbook

class Utils(softest.TestCase):
    def assetListItemText(self, list, value):

        for stop in list:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test passed")
            else:
                print("test failed")
        self.assert_all()


    def custom_logger(loglevel=logging.DEBUG):
        logger_name= inspect.stack()[1][3]
        logger= logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        fh= logging.FileHandler("automation.log")
        formatter = logging.Formatter("%(asctime)s - %(levelname)s : %(message)s ", datefmt="%m/%d/%Y %I:%M:%S %p")
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name_1, sheet):
        datalist = []
        wb = load_workbook(filename=file_name_1)
        sh = wb[sheet]

        row_ct = sh.max_row
        col_ct = sh.max_column
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(file_name_2):
        datalist = []
        csvdata= open(file_name_2, "r")
        reader_1 = csv.reader(csvdata)
        next(reader_1)
        for rows in reader_1:
            datalist.append(rows)
        return datalist